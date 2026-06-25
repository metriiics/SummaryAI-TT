from functools import lru_cache
from docling.document_converter import DocumentConverter
from pathlib import Path

import re

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from schemas import Documents, DocMetadata

class ParseDocs:
    """
        Класс для поддержки парсинга документов:
            doc, docx, xlsx, xls, pdf
    """
    def __init__(self):
        self.name_home_dir = "temp"
        self.abs_path = Path.cwd() / self.name_home_dir

        self.converter = DocumentConverter()

        self.parsers = {
            ".docx": self._parsing_docx_doc_pdf, 
            ".doc": self._parsing_docx_doc_pdf,
            ".pdf": self._parsing_docx_doc_pdf,
            ".xlsx": self._parsing_xlsx,
            ".xls": self._parsing_xlsx,
        }

        self.coef_tokens = 2.25
        self.PLACEHOLDER_RE = re.compile(r"^\{.*\}$")
    
    def _exist_home_dir(self):
        if not self.abs_path.exists():
            self.abs_path.mkdir()

    def _clean_txt(self):
        pass

    def _clean_xls(self, value) -> str:
        return " ".join(str(value).split()) if value is not None else ""

    def _check_file(self, file: str):
        path = self.abs_path / file
        if path.is_file():
            return path
        else:
            return "Файл не найден или не существует"

    def _detect_files(self, name_dir: str) -> list:
        self._exist_home_dir()

        dir = self.abs_path / name_dir
        if not dir.is_dir() or not dir.exists():
            return f"Директория не найдена или не существует: {dir}"

        files = []

        for file in dir.rglob("*"):
            if file.is_file():
                files.append(file)

        return files

    def router(self, name_dir: str):
        parse_content = []
        files = self._detect_files(name_dir)

        for file in files:
            if not file.is_file():
                continue

            parser = self.parsers.get(file.suffix.lower())
            if parser is None:
                print(f"Неизвестный формат: {file}")
                continue

            content = parser(file)
            parse_content.append(content)

        return parse_content

    def _parsing_docx_doc_pdf(self, file: str):
        path = self._check_file(file)

        result = self.converter.convert(path)
        doc = result.document
        markdown = doc.export_to_markdown()
        
        doc = Documents(
            name_docs=path.name,
            content=markdown,
            metadata=DocMetadata(
                file_type=path.suffix,
                length_char=len(markdown),
                length_word=len(markdown.split()),
                predicted_number_tokens=len(markdown.split()) * self.coef_tokens
            )
        )

        return doc

    def _parsing_xlsx(self, file: str):
        path = self._check_file(file)

        wb = load_workbook(path, data_only=True)

        lines = []

        for sheet in wb.worksheets:
            lines.extend([f"# Лист: {sheet.title}", ""])

            for row in sheet.iter_rows():

                values = [
                    text
                    for cell in row
                    if not isinstance(cell, MergedCell)
                    if (text := self._clean_xls(cell.value))
                    if not self.PLACEHOLDER_RE.fullmatch(text)
                ]

                if values:
                    lines.append(" | ".join(values))

            lines.extend(["", "-" * 80, ""])
 
        text = "\n".join(lines)
        doc = Documents(
            name_docs=path.name,
            content=text,
            metadata=DocMetadata(
                file_type=path.suffix,
                length_char=len(text),
                length_word=len(text.split()),
                predicted_number_tokens=len(text.split()) * self.coef_tokens
            )
        )

        return doc